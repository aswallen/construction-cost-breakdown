"""
Construction Cost Breakdown Automation Tool - FINAL VERSION

This script automates the process of extracting line items from builder documents
and populating a standardized Excel template while preserving ALL formatting.
FIXED: Extracts ALL rows AND preserves formatting by using openpyxl instead of pandas.
"""

import os
import json
import pandas as pd
from pathlib import Path
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False
    try:
        import PyPDF2
        PYPDF2_AVAILABLE = True
    except ImportError:
        PYPDF2_AVAILABLE = False

from PIL import Image
try:
    import pytesseract
    PYTESSERACT_AVAILABLE = True
except ImportError:
    PYTESSERACT_AVAILABLE = False

try:
    import google.generativeai as genai
    GOOGLE_AI_AVAILABLE = True
except ImportError:
    GOOGLE_AI_AVAILABLE = False
    genai = None

from typing import List, Dict, Union
import re
from openpyxl import load_workbook
from shutil import copy2

class ConstructionCostAutomation:
    def __init__(self, api_key: str = None):
        """
        Initialize the automation tool.
        
        Args:
            api_key: Google Gemini API key. If None, will try to get from environment.
        """
        self.api_key = api_key or os.getenv('GEMINI_API_KEY')
        
        if not GOOGLE_AI_AVAILABLE:
            print("Warning: Google Generative AI not available in this environment. AI parsing will not be available.")
            self.model = None
        elif self.api_key:
            genai.configure(api_key=self.api_key)
            self.model = genai.GenerativeModel('gemini-1.5-flash')
        else:
            print("Warning: No Gemini API key provided. AI parsing will not be available.")
            self.model = None
    
    def extract_text_from_file(self, file_path: str) -> str:
        """
        Extract text content from various file types.
        
        Args:
            file_path: Path to the file to extract text from
            
        Returns:
            Extracted text content as string
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_ext = file_path.suffix.lower()
        
        try:
            if file_ext == '.pdf':
                return self._extract_from_pdf(file_path)
            elif file_ext in ['.png', '.jpg', '.jpeg', '.bmp', '.tiff']:
                return self._extract_from_image(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                return self._extract_from_excel(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_ext}")
        except Exception as e:
            print(f"Error extracting text from {file_path}: {e}")
            return ""
    
    def _extract_from_pdf(self, file_path: Path) -> str:
        """Extract text from PDF file using available PDF library."""
        text_content = ""
        
        if PYMUPDF_AVAILABLE:
            # Use PyMuPDF (fitz) if available
            try:
                with fitz.open(file_path) as doc:
                    for page in doc:
                        text_content += page.get_text() + "\n"
                return text_content
            except Exception as e:
                print(f"PyMuPDF extraction failed: {e}")
        
        if PYPDF2_AVAILABLE:
            # Fallback to PyPDF2
            try:
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        text_content += page.extract_text() + "\n"
                return text_content
            except Exception as e:
                print(f"PyPDF2 extraction failed: {e}")
          # If no PDF library is available, return error message
        error_msg = "PDF extraction not available. Neither PyMuPDF nor PyPDF2 is installed."
        print(error_msg)
        raise ImportError(error_msg)
    
    def _extract_from_image(self, file_path: Path) -> str:
        """Extract text from image using OCR."""
        if not PYTESSERACT_AVAILABLE:
            error_msg = "OCR not available. Pytesseract is not installed or not available in this environment."
            print(error_msg)
            return f"[OCR NOT AVAILABLE] Image file: {file_path.name}. Please convert to PDF or use text-based documents for processing."
        
        try:
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            return text
        except Exception as e:
            print(f"OCR error: {e}")
            print("Note: Make sure Tesseract is installed and in your PATH")
            return f"[OCR FAILED] Image file: {file_path.name}. Error: {str(e)}"
    
    def _extract_from_excel(self, file_path: Path) -> str:
        """Extract text from Excel file."""
        try:
            # Try reading all sheets
            excel_file = pd.ExcelFile(file_path)
            all_text = ""
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                # Convert dataframe to string representation
                all_text += f"Sheet: {sheet_name}\n"
                all_text += df.to_string(na_rep='', index=False) + "\n\n"
            
            return all_text
        except Exception as e:
            print(f"Excel extraction error: {e}")
            return ""
    
    def parse_with_ai(self, text_content: str) -> List[Dict[str, Union[str, float]]]:
        """
        Use AI to parse text content and extract structured line items.
        
        Args:
            text_content: Raw text content from document
            
        Returns:
            List of dictionaries with 'line_item' and 'amount' keys
        """
        if not self.model:
            print("AI model not available. Please provide Gemini API key.")
            return []
        
        prompt = f"""
        You are an expert data entry assistant for construction cost breakdowns.
        Analyze the following text from a construction cost breakdown document.
        Your task is to identify each distinct line item and its corresponding monetary value.
        
        Instructions:
        1. Extract every line item and its cost from the document
        2. Exclude subtotals, grand totals, general contractor fees, or administrative fees
        3. Include only actual construction line items (permits, materials, labor, etc.)
        4. Format the output as a valid JSON array
        5. Each object should have exactly two keys: "line_item" and "amount"
        6. Ensure amounts are clean numbers (no currency symbols, commas, or text)
        7. If you see percentages or rates, convert them to actual dollar amounts if possible
        8. Line item names should be clear and descriptive
        
        Example format:
        [
            {{"line_item": "PERMITS", "amount": 10000.00}},
            {{"line_item": "EXCAVATION", "amount": 15000.00}},
            {{"line_item": "FOUNDATION", "amount": 38000.00}}
        ]
        
        Document text to analyze:
        ---
        {text_content}
        ---
        
        Return only the JSON array, no additional text or formatting.
        """
        
        try:
            response = self.model.generate_content(prompt)
            response_text = response.text.strip()
            
            # Clean the response - remove markdown formatting if present
            response_text = re.sub(r'```json\s*', '', response_text)
            response_text = re.sub(r'```\s*$', '', response_text)
            response_text = response_text.strip()
            
            # Parse JSON
            parsed_data = json.loads(response_text)
              # Validate the structure
            if isinstance(parsed_data, list):
                validated_data = []
                for item in parsed_data:
                    if isinstance(item, dict) and 'line_item' in item and 'amount' in item:
                        # Ensure amount is numeric
                        try:
                            amount = float(item['amount'])
                            validated_data.append({
                                'line_item': str(item['line_item']).strip(),
                                'amount': amount
                            })
                        except (ValueError, TypeError):
                            print(f"Skipping invalid amount: {item}")
                            continue
                return validated_data
            else:
                print("AI response is not a list format")
                return []
                
        except json.JSONDecodeError as e:
            print(f"Error parsing AI response as JSON: {e}")
            print(f"AI Response: {response_text[:500]}...")
            return []
        except Exception as e:
            print(f"Error during AI parsing: {e}")
            return []
    
    def populate_template(self, line_items: List[Dict], template_path: str, output_path: str) -> bool:
        """
        Populate Excel template while preserving ALL formatting, colors, and formulas.
        FIXED: Uses openpyxl + copy2 to preserve formatting while extracting all rows.
        
        Args:
            line_items: List of dictionaries with line_item and amount keys
            template_path: Path to the blank template Excel file
            output_path: Path where the populated file should be saved
            
        Returns:
            True if successful, False otherwise
        """
        try:
            print(f"📋 Starting template population...")
            print(f"📊 Processing {len(line_items)} line items")
            print(f"📂 Template: {template_path}")
            print(f"💾 Output: {output_path}")
            
            # STEP 1: Copy entire template to preserve ALL formatting
            copy2(template_path, output_path)
            print("✅ Template copied with all formatting preserved")
            
            # STEP 2: Open copied file keeping formulas intact
            workbook = load_workbook(output_path, data_only=False)  # Keep formulas!
            worksheet = workbook.active
            
            # STEP 3: Find column positions with improved logic
            line_item_col_idx = None
            amount_col_idx = None
            header_row = None
            
            print("🔍 Template analysis:")
            # Show template structure for debugging
            for row in range(1, 6):
                for col in range(1, 6):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        col_letter = worksheet.cell(row=row, column=col).column_letter
                        print(f"  {col_letter}{row}: {str(cell_value)[:50]}")
            
            # Look for line item column (prefer column A)
            print("🔍 Searching for Line Item column...")
            for row in range(1, 11):
                for col in range(1, 11):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        if "Line Item" in cell_value:
                            line_item_col_idx = col
                            header_row = row
                            col_letter = worksheet.cell(row=row, column=col).column_letter
                            print(f"🎯 Found Line Item column: {col_letter} (Col {col})")
                            break
                if line_item_col_idx:
                    break
            
            # Fallback search for line item column
            if line_item_col_idx is None:
                print("🔍 Trying broader Line Item search...")
                for row in range(1, 11):
                    for col in range(1, 6):  # Prefer early columns
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_lower = cell_value.lower()
                            if "line" in cell_lower and "item" in cell_lower:
                                line_item_col_idx = col
                                header_row = row
                                col_letter = worksheet.cell(row=row, column=col).column_letter
                                print(f"🎯 Found Line column: {cell_value} at {col_letter} (Col {col})")
                                break
                    if line_item_col_idx:
                        break
            
            # CRITICAL FIX: Always use column B for amounts (index 2)
            print("💰 FORCING amount column to B (Column 2)...")
            amount_col_idx = 2  # Column B - this guarantees column B usage!
            
            # Check what's in column B header for confirmation
            for row in range(1, 6):
                cell_value = worksheet.cell(row=row, column=2).value
                if cell_value and isinstance(cell_value, str):
                    print(f"💰 Column B header: '{cell_value}'")
                    break
            else:
                print("💰 Column B has no header, but using it anyway (as requested)")
            
            # Final fallback for line items
            if line_item_col_idx is None:
                print("⚠️ No line item column found, defaulting to Column A")
                line_item_col_idx = 1
            
            col_a_letter = worksheet.cell(row=1, column=line_item_col_idx).column_letter
            col_b_letter = worksheet.cell(row=1, column=amount_col_idx).column_letter
            print(f"📊 FINAL SETUP:")
            print(f"📊   Line Items → Column {col_a_letter} (index {line_item_col_idx})")
            print(f"📊   Amounts   → Column {col_b_letter} (index {amount_col_idx}) ← FORCED TO COLUMN B")
            
            # STEP 4: Find first empty row for data insertion
            insert_row = None
            start_row = (header_row + 1) if header_row else 2
            
            for row in range(start_row, worksheet.max_row + 20):
                cell_value = worksheet.cell(row=row, column=line_item_col_idx).value
                if cell_value is None or str(cell_value).strip() == '':
                    insert_row = row
                    break
            
            if insert_row is None:
                insert_row = worksheet.max_row + 1
            
            print(f"📍 Inserting {len(line_items)} items starting at row: {insert_row}")
            
            # STEP 5: Insert ALL the data while preserving cell formatting
            for i, item in enumerate(line_items):
                row_idx = insert_row + i
                
                # Insert line item
                line_item_cell = worksheet.cell(row=row_idx, column=line_item_col_idx)
                line_item_cell.value = item['line_item']
                
                # Insert amount as number for formulas - GUARANTEED TO BE COLUMN B
                amount_cell = worksheet.cell(row=row_idx, column=2)  # Hardcode column 2 = B
                amount_cell.value = float(item['amount'])
                
                # Show progress for first few and last few items
                if i < 5 or i >= len(line_items) - 2:
                    print(f"  ✅ Row {row_idx}: {item['line_item']} = ${item['amount']:,.2f}")
                    print(f"      Line Item → {col_a_letter}{row_idx}, Amount → B{row_idx} (FORCED)")
                elif i == 5:
                    print(f"      ... processing {len(line_items) - 7} more items ...")
            
            # STEP 6: Save preserving everything
            workbook.save(output_path)
            workbook.close()
            
            print(f"🎉 SUCCESS! File saved: {output_path}")
            print(f"📊 Successfully inserted {len(line_items)} line items total")
            print(f"✨ All formatting, colors, and formulas preserved!")
            print(f"💡 ALL AMOUNTS GUARANTEED TO BE IN COLUMN B!")
            return True
            
        except Exception as e:
            print(f"❌ Error populating template: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def process_document(self, input_file: str, template_path: str, output_path: str = None) -> str:
        """
        Complete workflow: extract text, parse with AI, and populate template.
        
        Args:
            input_file: Path to the builder's document
            template_path: Path to the blank template
            output_path: Optional output path. If None, will generate based on input filename
            
        Returns:
            Path to the generated file if successful, None otherwise
        """
        # Generate output path if not provided
        if output_path is None:
            input_path = Path(input_file)
            output_path = f"COMPLETED_{input_path.stem}_breakdown.xlsx"
        
        print(f"Processing document: {input_file}")
        
        # Step 1: Extract text
        print("Step 1: Extracting text from document...")
        extracted_text = self.extract_text_from_file(input_file)
        
        if not extracted_text.strip():
            print("No text could be extracted from the document.")
            return None
        
        print(f"Extracted {len(extracted_text)} characters of text")
        
        # Step 2: Parse with AI
        print("Step 2: Parsing with AI...")
        line_items = self.parse_with_ai(extracted_text)
        
        if not line_items:
            print("No line items could be extracted.")
            return None
        
        print(f"Extracted {len(line_items)} line items:")
        for item in line_items:
            print(f"  - {item['line_item']}: ${item['amount']:,.2f}")
        
        # Step 3: Populate template
        print("Step 3: Populating template...")
        success = self.populate_template(line_items, template_path, output_path)
        
        if success:
            return output_path
        else:
            return None


def main():
    """
    Main function to run the automation.
    You can modify this to suit your needs.
    """
    # Initialize the automation tool
    # Note: You'll need to set your Gemini API key
    automation = ConstructionCostAutomation()
    
    # Define paths
    template_path = "_Construction_Breakdown_Template_BLANK.xlsx"
    
    # Example documents to process
    example_docs = [
        "Cost break down-Prem.pdf",
        "Cost Breakdown -Constrution Arndt.pdf", 
        "Safranek cost.pdf"
    ]
    
    # Process each document
    for doc in example_docs:
        if os.path.exists(doc):
            print(f"\n{'='*50}")
            print(f"Processing: {doc}")
            print('='*50)
            
            result = automation.process_document(doc, template_path)
            
            if result:
                print(f"✅ Successfully processed {doc}")
                print(f"📄 Output saved to: {result}")
            else:
                print(f"❌ Failed to process {doc}")
        else:
            print(f"⚠️  File not found: {doc}")


if __name__ == "__main__":
    main()
